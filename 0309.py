# -*- coding: utf-8 -*-


"""
李运辰 2021-3-9

公众号：python爬虫数据分析挖掘
"""


import requests
from lxml import etree
import json
headers = {
            'user-agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3947.100 Safari/537.36',
        }


import openpyxl
outwb = openpyxl.Workbook()
outws = outwb.create_sheet(index=0)

outws.cell(row=1,column=1,value="index")
outws.cell(row=1,column=2,value="title")
outws.cell(row=1,column=3,value="price")
outws.cell(row=1,column=4,value="CommentCount")

count=2
###根据商品id获取评论数
def commentcount(product_id):
    url = "https://club.jd.com/comment/productCommentSummaries.action?referenceIds="+str(product_id)+"&callback=jQuery8827474&_=1615298058081"
    res = requests.get(url, headers=headers)
    res.encoding = 'gbk'
    text = (res.text).replace("jQuery8827474(","").replace(");","")
    text = json.loads(text)
    comment_count = text['CommentsCount'][0]['CommentCountStr']

    comment_count = comment_count.replace("+", "")
    ###对“万”进行操作
    if "万" in comment_count:
        comment_count = comment_count.replace("万","")
        comment_count = str(int(comment_count)*10000)


    return comment_count


#commentcount("71929438514")
###获取每一页的商品数据
def getlist(url):
    global  count
    #url="https://search.jd.com/search?keyword=笔记本&wq=笔记本&ev=exbrand_联想%5E&page=9&s=241&click=1"
    res = requests.get(url,headers=headers)
    res.encoding = 'utf-8'
    text = res.text


    selector = etree.HTML(text)
    list = selector.xpath('//*[@id="J_goodsList"]/ul/li')

    for i in list:
        title=i.xpath('.//div[@class="p-name p-name-type-2"]/a/em/text()')[0]
        price = i.xpath('.//div[@class="p-price"]/strong/i/text()')[0]
        product_id = i.xpath('.//div[@class="p-commit"]/strong/a/@id')[0].replace("J_comment_","")

        comment_count = commentcount(product_id)
        #print(title)
        #print(price)
        #print(comment_count)

        outws.cell(row=count, column=1, value=str(count-1))
        outws.cell(row=count, column=2, value=str(title))
        outws.cell(row=count, column=3, value=str(price))
        outws.cell(row=count, column=4, value=str(comment_count))

        count = count +1
        #print("-----")


#遍历每一页
def getpage():
    page=1
    s = 1
    for i in range(1,6):
        print("page="+str(page)+",s="+str(s))
        url = "https://search.jd.com/search?keyword=笔记本&wq=笔记本&ev=exbrand_联想%5E&page="+str(page)+"&s="+str(s)+"&click=1"
        getlist(url)
        page = page+2
        s = s+60



#开始爬取
getpage()
#getlist()

outwb.save("京东商品-李运辰.xls")#保存


"""
1.入门爬虫（京东商品数据为例）
2.如何获取网页标签内容
3.获取评论数
4.如何通过python将数据保存到excel
5.遍历下一页


源码获取方式：
公众号：python爬虫数据分析挖掘

回复：京东商品

"""